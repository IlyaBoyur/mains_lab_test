from datetime import datetime
from typing import Any, Dict, List

import openpyxl
from config.settings import DATA_DIR


def parse_excel(xlsx_file: str) -> List[Dict[str, Any]]:
    """Retrieve list of items constructed from excel table row."""
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.active
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)
    items = []
    for row in range(2, sheet.max_row+1):
        item = {}
        for column in range(1, sheet.max_column):
            item[col_names[column-1]] = sheet.cell(row=row,
                                                   column=column).value
        items.append(item)
    return items


def get_bills_list(
    xlsx: str = f'{DATA_DIR}/bills.xlsx'
) -> List[Dict[str, Any]]:
    """Retrieve and filter bills"""
    bills = parse_excel(f'{DATA_DIR}/bills.xlsx')
    return [bill for bill in bills
            if (
                bill['service'] != '-'
                and isinstance(bill['date'], datetime)
                and isinstance(bill['№'], int)
            )]
